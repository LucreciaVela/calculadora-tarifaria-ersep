
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
import os

# Diccionario de descripciones para campos de ingreso
descripcion_variables = {
    "MT": "Monto anual de la prima para personal de conducción",
    "U": "Costo de los uniformes según convenio",
    "Nc": "Valuación neumático 900 r22.5",
    "Nm": "Valuación neumático 275/80 r22.5",
    "Ng": "Valuación neumático 295/80 r22.5",
    "L": "Costo de lubricantes anualizado",
    "Mp": "Mantenimiento preventivo anual",
    "E": "Costo de equipos electrónicos",
    "Pp": "Costo promedio de patentes",
    "RTM": "Renovación técnica y municipal",
    "Sbcu": "Sueldo básico conductor único",
    "Pm": "Seguros varios mensuales",
    "Pr": "Reparaciones generales mensuales",
    "SBcg": "Sueldo básico conductor general",
    "Gm": "Gastos de mantenimiento anual",
    "Gr": "Gastos de reparación anual",
    "Vc": "Valor compra vehículo tipo C",
    "Vm": "Valor compra vehículo tipo M",
    "Vg": "Valor compra vehículo tipo G",
    "RBM": "Remuneración bruta mensual promedio",
    "Ut": "Uso total anual estimado"
}

# Encabezado automático
meses_es = {1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO", 7: "JULIO",
            8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"}
hoy = datetime.now()
encabezado = f"CÁLCULO TARIFARIO A {meses_es[hoy.month]} {hoy.year}"

st.set_page_config(page_title="ERSeP Transporte", layout="wide")
st.title("🚌 Calculadora Tarifaria ERSeP Transporte")
st.markdown(f"### {encabezado}")
st.markdown("#### Ingreso de datos variables")

EXCEL_FILE = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"
if not os.path.exists(EXCEL_FILE):
    st.error("❌ No se encuentra el archivo base.")
    st.stop()

# Cargar valores de referencia desde columna O
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb["Hoja Llave"]
valores_referencia = {}
for row in range(2, ws.max_row + 1):
    cod = ws.cell(row=row, column=1).value
    ref = ws.cell(row=row, column=15).value
    if cod in descripcion_variables:
        valores_referencia[cod] = float(ref) if isinstance(ref, (int, float)) else 0.0

# Mostrar formulario
inputs = {}
cols = st.columns(3)
for i, cod in enumerate(descripcion_variables):
    label = f"{cod} – {descripcion_variables[cod]}"
    with cols[i % 3]:
        inputs[cod] = st.number_input(label, value=valores_referencia.get(cod, 0.0), format="%.2f", step=100.0)

if st.button("🔍 Calcular y mostrar resultados"):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Hoja Llave"]
    for row in range(2, ws.max_row + 1):
        cod = ws.cell(row=row, column=1).value
        if cod in inputs:
            ws.cell(row=row, column=2).value = inputs[cod]
    wb.active = wb["Resumen de Calculo"]
    wb.save("archivo_actualizado.xlsx")

    df_raw = pd.read_excel("archivo_actualizado.xlsx", sheet_name="Resumen de Calculo", header=None)

    # Buscar cualquier celda que contenga "CALCULO TARIFARIO"
    fila_inicio = None
    for i in range(len(df_raw)):
        for j in range(len(df_raw.columns)):
            val = str(df_raw.iloc[i, j])
            if "CALCULO TARIFARIO" in val.upper():
                fila_inicio = i
                break
        if fila_inicio is not None:
            break

    if fila_inicio is None:
        st.error("❌ No se encontró el encabezado de la tabla principal.")
        st.stop()

    df = df_raw.iloc[fila_inicio + 2:, :6]
    df.columns = ["Código", "Ítem de costo", "Valor calculado", "% Incidencia", "Valor vigente", "Variación %"]
    df = df.dropna(subset=["Código", "Ítem de costo"], how="all")

    # Reemplazar subtotales
    reemplazos = {
        "A8": "TOTAL COSTOS ASOCIADOS AL PERSONAL",
        "B6": "TOTAL COSTOS VARIABLES ASOCIADOS AL VEHÍCULO",
        "C5": "TOTAL COSTOS FIJOS ASOCIADOS AL VEHÍCULO",
        "D7": "TOTAL COSTOS EMPRESARIOS E IMPOSITIVOS",
        "Z1": "COSTOS MEDIOS PRESUPUESTADOS"
    }
    df["Ítem de costo"] = df.apply(
        lambda row: reemplazos[row["Código"]] if row["Código"] in reemplazos else row["Ítem de costo"], axis=1
    )

    def fmt_decimal(val):
        try:
            return f"{float(val):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except:
            return val

    def fmt_pct(val):
        try:
            return f"{float(val):.1f}".replace(".", ",") + " %"
        except:
            return val

    df["Valor calculado"] = df["Valor calculado"].apply(fmt_decimal)
    df["% Incidencia"] = df["% Incidencia"].apply(fmt_pct)
    df["Valor vigente"] = df["Valor vigente"].apply(fmt_decimal)
    df["Variación %"] = df["Variación %"].apply(fmt_pct)

    st.success("✅ Cálculo completado correctamente.")
    st.dataframe(df, use_container_width=True, hide_index=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Resumen")
        wb = writer.book
        ws = writer.sheets["Resumen"]
        formato = wb.add_format({"border": 1})
        for col in range(len(df.columns)):
            ws.set_column(col, col, 28, formato)

    st.download_button(
        label="📥 Descargar Resumen con diseño",
        data=output.getvalue(),
        file_name="Resumen_Tarifario_ERSEP_OK.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
