
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime

meses_es = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}
hoy = datetime.now()
encabezado = f"CÁLCULO TARIFARIO A {meses_es[hoy.month]} {hoy.year}"

st.set_page_config(page_title="Calculadora Tarifaria ERSeP Transporte", layout="wide")
st.title("🚌 Calculadora Tarifaria ERSeP Transporte")
st.subheader("Ingreso de datos variables")

EXCEL_FILE = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"

variables = [
    {"codigo": "MT", "descripcion": "Monto anual de la prima para personal de conducción", "valor": 451812.74},
    {"codigo": "U", "descripcion": "Costo de los uniformes según convenio. Por temporada", "valor": 188544.80},
    {"codigo": "Nc", "descripcion": "Un. de valuación (medidas 900 r22,5) + 0,6 (1 recapado)", "valor": 530495.80},
    {"codigo": "Nm", "descripcion": "Un. de valuación (medidas 275/80 r22,5) + 0,6 (1 recapado)", "valor": 530495.80},
    {"codigo": "Ng", "descripcion": "Un. de valuación (medidas 295/80 r22,5) + 0,6 (1 recapado)", "valor": 545795.30},
    {"codigo": "L", "descripcion": "Costo lubricantes anualizado", "valor": 150000.00},
    {"codigo": "Mp", "descripcion": "Costo medio por mantenimiento preventivo anual", "valor": 225000.00},
    {"codigo": "E", "descripcion": "Costo de equipo (GPS, ticketera, etc.) anualizado", "valor": 84500.00},
    {"codigo": "Pp", "descripcion": "Costo promedio de patentes anual", "valor": 95600.00},
    {"codigo": "RTM", "descripcion": "Costo de renovación técnica y municipal", "valor": 64789.00},
    {"codigo": "Sbcu", "descripcion": "Sueldo básico del conductor único", "valor": 800000.00},
    {"codigo": "Pm", "descripcion": "Costo promedio mensual de seguros varios", "valor": 120000.00},
    {"codigo": "Pr", "descripcion": "Costo promedio mensual de reparaciones generales", "valor": 200000.00},
    {"codigo": "SBcg", "descripcion": "Sueldo básico del conductor general", "valor": 780000.00},
    {"codigo": "Gm", "descripcion": "Gastos de mantenimiento anual", "valor": 265000.00},
    {"codigo": "Gr", "descripcion": "Gastos de reparación anual", "valor": 180000.00},
    {"codigo": "Vc", "descripcion": "Valor de compra del vehículo tipo C", "valor": 7500000.00},
    {"codigo": "Vm", "descripcion": "Valor de compra del vehículo tipo M", "valor": 8750000.00},
    {"codigo": "Vg", "descripcion": "Valor de compra del vehículo tipo G", "valor": 9200000.00},
    {"codigo": "RBM", "descripcion": "Remuneración bruta mensual promedio", "valor": 900000.00},
    {"codigo": "Ut", "descripcion": "Uso total anual estimado del vehículo", "valor": 120000.00},
]

user_inputs = {}
cols = st.columns(3)
for i, var in enumerate(variables):
    with cols[i % 3]:
        val = st.number_input(
            f"{var['codigo']} – {var['descripcion']}",
            value=round(var["valor"], 2),
            format="%.2f",
            step=100.0
        )
        user_inputs[var["codigo"]] = val

if st.button("🔍 Calcular y mostrar resultados"):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        hoja_llave = wb["Hoja Llave"]
        for row in range(2, hoja_llave.max_row + 1):
            codigo = hoja_llave.cell(row=row, column=1).value
            if isinstance(codigo, str) and codigo.strip() in user_inputs:
                hoja_llave.cell(row=row, column=2).value = user_inputs[codigo.strip()]
        wb.active = wb["Resumen de Calculo"]
        wb.save("actualizado.xlsx")

        df = pd.read_excel("actualizado.xlsx", sheet_name="Resumen de Calculo")
        df = df.dropna(how="all")

        st.markdown(f"### {encabezado}")
        st.dataframe(df, use_container_width=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Resumen')
        st.download_button(
            label="📥 Descargar Resumen en Excel",
            data=output.getvalue(),
            file_name="Resumen_Tarifario_ERSEP.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"❌ Error al procesar el cálculo: {e}")
