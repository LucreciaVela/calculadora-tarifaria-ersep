
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime

# Encabezado automático
meses_es = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}
hoy = datetime.now()
encabezado = f"CÁLCULO TARIFARIO A {meses_es[hoy.month]} {hoy.year}"

st.set_page_config(page_title="Calculadora ERSeP Transporte", layout="wide")
st.title("🚌 Calculadora Tarifaria ERSeP Transporte")
st.markdown(f"### {encabezado}")
st.markdown("**Ingreso de datos variables**")

# Cargar archivo base
EXCEL_FILE = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"

# Lista de variables esperadas
variables = [
    "MT", "U", "Nc", "Nm", "Ng", "L", "Mp", "E", "Pp",
    "RTM", "Sbcu", "Pm", "Pr", "SBcg", "Gm", "Gr",
    "Vc", "Vm", "Vg", "RBM", "Ut"
]

# Cargar valores de referencia desde la columna O
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb["Hoja Llave"]
valores_referencia = {}
for row in range(2, ws.max_row + 1):
    cod = ws.cell(row=row, column=1).value
    ref = ws.cell(row=row, column=15).value
    if cod in variables:
        valores_referencia[cod] = round(ref, 2) if isinstance(ref, (float, int)) else 0.0

# Formulario
inputs = {}
cols = st.columns(3)
for i, cod in enumerate(variables):
    with cols[i % 3]:
        inputs[cod] = st.number_input(f"{cod}", value=valores_referencia.get(cod, 0.0), format="%.2f")

# Al presionar calcular
if st.button("🔍 Calcular y mostrar resultados"):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Hoja Llave"]
    for row in range(2, ws.max_row + 1):
        cod = ws.cell(row=row, column=1).value
        if cod in inputs:
            ws.cell(row=row, column=2).value = inputs[cod]
    wb.save("archivo_actualizado.xlsx")

    df = pd.read_excel("archivo_actualizado.xlsx", sheet_name="Resumen de Calculo")
    df = df.dropna(how="all")

    st.success("✅ Cálculo completado correctamente.")
    st.dataframe(df, use_container_width=True)

    # Descargar versión Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Resumen")
    st.download_button(
        label="📥 Descargar Resumen en Excel",
        data=output.getvalue(),
        file_name="Resumen_Tarifario_ERSEP.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
