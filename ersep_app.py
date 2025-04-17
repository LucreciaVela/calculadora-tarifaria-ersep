
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

# Cargar archivo base
EXCEL_FILE = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"

st.set_page_config(page_title="Calculadora Tarifaria ERSeP Transporte", layout="centered")

st.title("🚌 Calculadora Tarifaria ERSeP Transporte")

st.subheader("📥 Ingreso de Datos (últimos valores de referencia visibles)")

# Datos variables extraídos previamente
variables = [
    {"codigo": "MT", "descripcion": "Monto anual de la prima para personal de conducción", "valor": 451812.74},
    {"codigo": "U", "descripcion": "Costo de los uniformes según convenio. Por temporada", "valor": 188544.8},
    {"codigo": "Nc", "descripcion": "Un. de valuación (medidas 900 r22,5) + 0,6 (1 recapado)", "valor": 530495.798},
    {"codigo": "Nm", "descripcion": "Un. de valuación (medidas 275/80 r22,5) + 0,6 (1 recapado)", "valor": 530495.798},
    {"codigo": "Ng", "descripcion": "Un. de valuación (medidas 295/80 r22,5) + 0,6 (1 recapado)", "valor": 545795.298},
]

# Crear diccionario de entradas
user_inputs = {}

# Mostrar los campos
for var in variables:
    val = st.number_input(
        f"{var['codigo']} – {var['descripcion']}",
        value=var["valor"],
        format="%.3f",
        step=1.0
    )
    user_inputs[var["codigo"]] = val

# Procesar cálculo y mostrar resultados
if st.button("🔍 Calcular y mostrar resultados"):
    try:
        # Abrir Excel con fórmulas
        wb = openpyxl.load_workbook(EXCEL_FILE)
        hoja_llave = wb["Hoja Llave"]

        # Insertar los valores modificados en sus respectivas celdas
        codigo_col_map = {}
        for row in range(2, hoja_llave.max_row + 1):
            codigo = hoja_llave.cell(row=row, column=1).value
            if isinstance(codigo, str) and codigo.strip() in user_inputs:
                hoja_llave.cell(row=row, column=2).value = user_inputs[codigo.strip()]

        # Calcular
        wb.active = wb["Resumen de Calculo"]
        wb.save("archivo_actualizado.xlsx")

        # Leer hoja para mostrar
        df = pd.read_excel("archivo_actualizado.xlsx", sheet_name="Resumen de Calculo")
        df = df.dropna(how="all")  # eliminar filas completamente vacías
        st.success("✅ Cálculo realizado correctamente")
        st.dataframe(df)

        # Exportar la misma tabla a Excel para descarga
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Resumen')
        st.download_button(
            label="📥 Descargar Resumen en Excel",
            data=output.getvalue(),
            file_name="Resumen_Tarifario_ERSEP.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Error al procesar el cálculo: {e}")
