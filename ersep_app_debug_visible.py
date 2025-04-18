
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
import os

# Encabezado
meses_es = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}
hoy = datetime.now()
encabezado = f"CÁLCULO TARIFARIO A {meses_es[hoy.month]} {hoy.year}"

st.set_page_config(page_title="DEBUG ERSeP Transporte", layout="wide")
st.title("🧪 DEBUG Calculadora ERSeP Transporte")
st.markdown(f"### {encabezado}")

EXCEL_FILE = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"

# Verificar existencia del archivo
if not os.path.exists(EXCEL_FILE):
    st.error(f"❌ Archivo '{EXCEL_FILE}' no encontrado. Asegurate de que esté en la misma carpeta que este script.")
    st.stop()

try:
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_llave = wb["Hoja Llave"]
    ws_resumen = wb["Resumen de Calculo"]
except Exception as e:
    st.error(f"❌ Error al abrir hojas del archivo: {e}")
    st.stop()

# Mostrar éxito de apertura
st.success("✅ Archivo Excel cargado y hojas encontradas correctamente.")

# Verificar primeras celdas clave
try:
    codigo_ejemplo = ws_llave.cell(row=2, column=1).value
    valor_ejemplo = ws_llave.cell(row=2, column=2).value
    st.info(f"Primer código en Hoja Llave: **{codigo_ejemplo}** – Valor: **{valor_ejemplo}**")
except Exception as e:
    st.error(f"❌ No se pudieron leer celdas clave de la hoja 'Hoja Llave': {e}")
    st.stop()

# Intentar leer tabla resumen
try:
    df_resumen = pd.read_excel(EXCEL_FILE, sheet_name="Resumen de Calculo")
    st.success("✅ Hoja 'Resumen de Calculo' cargada correctamente.")
    st.dataframe(df_resumen.head(10))
except Exception as e:
    st.error(f"❌ Error al leer la hoja 'Resumen de Calculo': {e}")
