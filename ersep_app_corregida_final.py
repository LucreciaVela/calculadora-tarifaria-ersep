
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime

# Mes y año automático
meses_es = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}
hoy = datetime.now()
encabezado = f"CÁLCULO TARIFARIO A {meses_es[hoy.month]} {hoy.year}"

st.set_page_config(page_title="Calculadora ERSeP Transporte", layout="wide")
st.title("🚌 Calculadora Tarifaria ERSeP Transporte")
st.markdown(f"### {encabezado}")
st.markdown("**Ingreso de datos variables**")

# Archivo base
EXCEL_FILE = "Incremento TBK_ Mesa 13 Octubre 2025.xlsx"

# Diccionario con descripciones
descripcion_variables = {
    "MT": "Monto anual de la prima para personal de conducción",
    "U": "Costo de los uniformes según convenio",
    "Nc": "Valuación neumático 900 r22.5",
    "Nm": "Valuación neumático 275/80 r22.5",
    "Ng": "Valuación neumático 295/80 r22.5",
    "L": "Costo de lubricantes anualizado",
    "Mp": "Mantenimiento preventivo anual",
    "E": "Costo de equipos electrónicos",
    "Pp": "Costo promedio de patentes",
    "RTM": "Renovación técnica y municipal",
    "Sbcu": "Sueldo básico conductor único",
    "Pm": "Seguros varios mensuales",
    "Pr": "Reparaciones generales mensuales",
    "SBcg": "Sueldo básico conductor general",
    "Gm": "Gastos de mantenimiento anual",
    "Gr": "Gastos de reparación anual",
    "Vc": "Valor compra vehículo tipo C",
    "Vm": "Valor compra vehículo tipo M",
    "Vg": "Valor compra vehículo tipo G",
    "RBM": "Remuneración bruta mensual promedio",
    "Ut": "Uso total anual estimado"
}

# Cargar referencias desde columna O
wb_ref = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws_ref = wb_ref["Hoja Llave"]
valores_referencia = {}
for row in range(2, ws_ref.max_row + 1):
    cod = ws_ref.cell(row=row, column=1).value
    ref = ws_ref.cell(row=row, column=15).value
    if cod in descripcion_variables:
        valores_referencia[cod] = float(ref) if isinstance(ref, (int, float)) else 0.0

# Formulario con descripciones
inputs = {}
cols = st.columns(3)
for i, cod in enumerate(descripcion_variables):
    label = f"{cod} – {descripcion_variables[cod]}"
    with cols[i % 3]:
        valor = float(valores_referencia.get(cod, 0.0))
        inputs[cod] = st.number_input(label, value=valor, format="%.2f", step=100.0)

# Al calcular
if st.button("🔍 Calcular y mostrar resultados"):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Hoja Llave"]
    for row in range(2, ws.max_row + 1):
        cod = ws.cell(row=row, column=1).value
        if cod in inputs:
            ws.cell(row=row, column=2).value = inputs[cod]
    wb.save("archivo_actualizado.xlsx")

    df_raw = pd.read_excel("archivo_actualizado.xlsx", sheet_name="Resumen de Calculo")
    df_raw = df_raw.dropna(how="all")

    # Procesar tabla para dejar solo las 6 columnas necesarias
    columnas_finales = ["Código", "Ítem de costo", "Valor calculado", "% Incidencia", "Valor vigente", "Variación %"]
    df_final = pd.DataFrame(columns=columnas_finales)

    for index, row in df_raw.iterrows():
        try:
            cod = str(row[0])
            item = str(row[1])
            val_nuevo = float(row[2])
            incidencia = float(row[3])
            val_viejo = float(row[4])
            variacion = float(row[5])
            df_final.loc[len(df_final)] = [
                cod,
                item,
                f"{val_nuevo:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"{incidencia:.1f}".replace(".", ",") + " %",
                f"{val_viejo:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"{variacion:.1f}".replace(".", ",") + " %"
            ]
        except:
            continue

    st.success("✅ Cálculo completado correctamente.")
    st.dataframe(df_final, use_container_width=True)

    # Descargar como Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Resumen")
        wb = writer.book
        ws = writer.sheets["Resumen"]
        formato = wb.add_format({'border': 1})
        for col in range(6):
            ws.set_column(col, col, 28, formato)

    st.download_button(
        label="📥 Descargar Resumen con Diseño",
        data=output.getvalue(),
        file_name="Resumen_Tarifario_ERSEP.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
